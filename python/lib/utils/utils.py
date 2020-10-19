import os
import pandas as pd
import numpy as np

import openpyxl
from openpyxl.drawing.image import Image

import pickle


def read_pickle(pkl_path: str):
    """pickle データをロードする関数

    Param
    -----
    pkl_path (str):
        読み込む pickle データのパス

    Return
    ------
    data:
        pickle から読み込んだデータ
    """
    with open(pkl_path, 'rb') as f:
        return pickle.load(f)


def image_past(worksheet, cell_name: str, img_path: str):
    """Excel に画像を貼りつける関数

    Parames
    -------
    worksheet (openpyxl.Worksheet):
        画像を貼り付け先となる Excel の worksheet
    cell_name (str):
        画像を貼り付ける cell の番号
        例 "C2"
    img_path (str):
        貼り付ける画像をロードするためのパス
    """

    try:
        # 画像ファイルのロード
        img = Image(img_path)
        # 画像を貼り付ける
        worksheet.add_image(img, cell_name)
    except FileNotFoundError:
        # 画像が見つからない場合はスキップ
        pass


def image_past_AllColumn(worksheet, past_column: str):
    """Excel の縦列に画像を貼り付ける関数

    Parames
    -------
    worksheet (openpyxl.Worksheet):
        画像を貼り付け先となる Excel の worksheet
    past_column (str):
        画像を貼り付ける列名
    """

    # 列数を取得
    row_max = len(list(worksheet.rows))
    # 画像を貼り付け
    for i in range(1, row_max):
        idx = i + 1
        cell_name = past_column + str(idx)  # "C"+ "1" → "C1"
        img_path = worksheet[cell_name].value

        image_past(worksheet, cell_name, img_path)


def save_Excel(path: str, data: dict, img_past: bool = False):
    """辞書型のデータを Excel ファイル年て保存する関数

    Params
    ------
    path (str):
        保存する Excel ファイルへのパス
    data (dict):
        保存するデータ
    img_past (bool):
        True: 画像をセル内の画像パスと置き換える。
        False: 画像を画像パスと置き換えない。
    """

    # 辞書の値を list 型に変換
    values = [data.values()]
    # pandas DataFrame にデータを変換
    # 最初の列に key を表示
    df = pd.DataFrame(values, columns=data.keys())
    # .xlsx ファイルとして保存
    df.to_excel(path, index=False)

    if img_past:
        # Excel のロード
        wb = openpyxl.load_workbook(path)
        # Active シート取得
        ws = wb.active
        print(type(ws))
        # 画像パスの cell を画像に置き換える
        # 画像パスが入っている cell の列を取得
        # そのために、最初の行を走査する。
        col_max = ws.max_column  # 列数の最大値を取得

        for i in range(1, col_max):
            cell_value = ws.cell(column=i, row=1).value
            if 'image_name' in cell_value:
                past_column = openpyxl.utils.get_column_letter(
                    ws.cell(column=i, row=1).column)

                image_past_AllColumn(ws, past_column)

        wb.save(path)


if __name__ == "__main__":
    path = os.path.dirname(os.path.abspath(__file__))

    data = {
        'object_name': 'ape',
        'local_idx': 1,
        'image_name': 'color1.jpg',
    }

    save_Excel(path, data)
